import csv
import re
import sys
from typing import List, Dict
from openpyxl import load_workbook

# match section headers that typically introduce excipient lists
LABEL_PAT = re.compile(
    r"\b(inactive ingredients?|inactives?|other ingredients|inactive components|nonmedicinal ingredients|preservatives?|inert ingredients)\b",
    re.I,
)
# remove numbers followed by common concentration units (mg, g, %, etc.)
UNIT_PAT = re.compile(r"\b\d+(?:\.\d+)?\s*(mg|g|kg|mcg|ug|µg|ml|l|%)\b", re.I)

# tokens containing any of these keywords will be moved to the notes column
NOTE_KEYWORDS = {
    "capsule",
    "capsules",
    "tablet",
    "tablets",
    "structural",
    "formula",
    "vial",
    "bottle",
    "oral",
    "imprinted",
    "available",
    "granules",
    "reconstituted",
    "diluted",
    "administration",
    "inactive",
    "constituted",
    "form",
    "serotonin",
    "reuptake",
    "inhibitor",
    "image",
    "active",
    "actives",
    "coating",
    "system",
    "components",
    "performance",
    "mixture",
    "racemic",
    "syringe",
    "needle",
    "plunger",
    "stopper",
    "kit",
    "cap",
    "backstop",
    "rod",
    "safety",
    "walled",
}


def read_excel(path: str) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) if h is not None else f"col{idx}" for idx, h in enumerate(rows[0])]
    data: List[Dict[str, str]] = []
    for row in rows[1:]:
        row_dict: Dict[str, str] = {}
        for h, v in zip(header, row):
            row_dict[h] = str(v) if v is not None else ""
        data.append(row_dict)
    return data


def clean_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r"\([^)]*\)", "", text)
    text = text.replace("\n", " ")
    text = re.sub(r"\bno\.\s*(\d+)", r"no \1", text, flags=re.I)
    text = text.replace("and/or", "and or")
    text = re.sub(r"\s+", " ", text)
    text = UNIT_PAT.sub("", text)
    text = text.replace(".", ";")
    text = re.sub(r"[^a-z0-9,;\s-]", " ", text.lower())
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _remove_leading(seg: str) -> str:
    seg = seg.lstrip()
    colon = seg.find(":")
    verb = re.search(r"\b(are|include|contain|consist of)\b", seg, re.I)
    if colon != -1 and (verb is None or colon < verb.start()):
        seg = seg[colon + 1 :]
    elif verb:
        seg = seg[verb.end() :]
    return seg.lstrip(" :;,")


def parse_from_description(desc: str) -> str:
    if not desc:
        return ""
    matches = list(LABEL_PAT.finditer(desc))
    if not matches:
        return ""
    segments = []
    for idx, m in enumerate(matches):
        start = m.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(desc)
        seg = desc[start:end]
        segments.append(_remove_leading(seg))
    text = " ".join(segments)
    text = re.sub(
        r"(structural formula|chemical structure|chem_structure|image of).*",
        "",
        text,
        flags=re.I,
    )
    text = re.sub(r"\s+is the (?:coloring agent|ink pigment)[^.;]*", "", text, flags=re.I)
    text = re.sub(r"\bthe tablet coating (?:consists of|is composed of)\b", "", text, flags=re.I)
    text = re.sub(r"and the following colorants?", ";", text, flags=re.I)
    text = re.split(r"system components", text, flags=re.I)[0]
    return text.strip()


def split_excipients_notes(text: str):
    raw_tokens = [p.strip() for p in re.split(r"[;,]", text) if p.strip()]
    tokens = []
    for p in raw_tokens:
        if re.search(r"\b(and|or)\b", p) and not re.search(r"\d", p):
            parts = [x.strip() for x in re.split(r"\b(?:and|or)\b", p) if x.strip()]
            tokens.extend(parts)
        else:
            tokens.append(p)
    excipients = []
    notes = []
    for token in tokens:
        if not re.search(r"[a-z]", token):
            continue
        m = re.search(r"\bcontains?\b", token)
        if m:
            before = token[: m.start()].strip()
            after = token[m.end() :].strip()
            if before:
                notes.append(before)
            token = after
        token = re.sub(r"^(and|or)\s+", "", token)
        if any(k in token for k in NOTE_KEYWORDS) or (
            "suspension" in token and len(token.split()) > 2
        ) or re.search(r"oral\s+solution", token):
            notes.append(token)
        else:
            excipients.append(token)
    # dedupe while preserving order
    def dedupe(items):
        seen = set()
        result = []
        for item in items:
            if item not in seen:
                seen.add(item)
                result.append(item)
        return result

    return dedupe(excipients), dedupe(notes)


def main():
    infile = sys.argv[1] if len(sys.argv) > 1 else "input.xlsx"
    outfile = sys.argv[2] if len(sys.argv) > 2 else "drug_excipients.csv"
    rows = read_excel(infile)
    results = []
    for row in rows:
        candidates = [
            row.get("English Product Name", ""),
            row.get("English Common Name", ""),
            row.get("English Drug Name", ""),
        ]
        product = ""
        for cand in candidates:
            if cand and re.search(r"[A-Za-z]", cand):
                product = cand.strip()
                break
        if not product:
            product = candidates[0]

        desc = row.get("Drug Description", "") or row.get("English Description", "")
        excip = parse_from_description(desc)
        if not excip:
            excip = row.get("Excipients", "")
        cleaned = clean_text(excip)
        if not cleaned:
            results.append((product, "", ""))
            continue
        excipients, notes = split_excipients_notes(cleaned)
        product_keywords = [c.lower() for c in candidates if c]
        filtered_excipient_list = []
        for e in excipients:
            if any(pk in e for pk in product_keywords):
                notes.append(e)
            else:
                filtered_excipient_list.append(e)
        results.append((product, "; ".join(filtered_excipient_list), "; ".join(notes)))

    with open(outfile, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["product", "excipients", "notes"])
        for product, excips, notes in results:
            writer.writerow([product, excips, notes])


if __name__ == "__main__":
    main()
